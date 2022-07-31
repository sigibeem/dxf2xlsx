import openpyxl
wb = openpyxl.Workbook()
ws = wb.active

list12 = [1, 2, 3, 4, 1, 2, 3, 4, 1, 2, 3, 4]
k = 1
col = 1
listIndex = []
for i in range(len(list12)):
    if list12[i] == 1:
        listIndex.append(i)
for j in range(len(list12)):
    if j in listIndex and j != 0:
        k += 1
        col =1
        wt = list12[j]
        ws.cell(k, col, list12[j])
        col += 1
    else:
        ws.cell(k, col, list12[j])
        wb2 = list12[j]
        col += 1

wb.save("C:\\Users\\田島\\desktop\\dad.xlsx")