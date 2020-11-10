#agregate_sales mine.py学習中に，セルに入力する際に文字列の場合""が必要なのか調べた
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "おほ"
"""
for a in range(len(list)):
    ws["A + a"].value = list[a]
"""
